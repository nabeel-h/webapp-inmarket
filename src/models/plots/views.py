from flask import Blueprint, request, render_template, flash
from werkzeug.utils import redirect, secure_filename
import pygal
from pygal.style import DarkStyle
from openpyxl import load_workbook
import os

from src.models.sheets.analysis import get_spreadsheetID, return_prepped_dfs_from_gsheets, return_prepped_dfs_from_excel, potential_sales_df
import src.models.sheets.errors as SheetErrors

__author__ = "nblhn"

plot_blueprint = Blueprint('plots', __name__)

ALLOWED_EXTENSIONS = set(['xlsx'])

@plot_blueprint.route('/g_plots', methods=['POST', 'GET'])
def create_gplots():
    '''
    if request.method == 'POST':
        spreadsheet_url = request.form['url']

        # check if Sheets ID from URL is valid
        spreadsheetID = get_spreadsheetID(spreadsheet_url)
        if spreadsheetID:
            # Checks if Named ranges exist in the Spreadsheet
            try:
                x = return_prepped_dfs_from_gsheets(spreadsheetID)
                ages_df = x[0]
                chain_df = x[1]
                stores_df = x[2]
                eating_places_df = x[3]
                restaurant_type_df = x[4]

                age_graph = create_age_plot(ages_df)
                category_graph = create_category_plot(chain_df)
                store_graph = create_stores_plot(stores_df)
                restaurant_summary_graph = create_restaurant_summary_plot(restaurant_type_df)
                restaurants_graph = create_restaurant_plot(eating_places_df)

                potential_sales = potential_sales_df(stores_df, chain_df)
                total_potential_customers = potential_sales["Potential Additional Customers"].sum()

                return render_template('plots/plots.jinja2', chain_df=chain_df, stores_df=stores_df, ages_df=ages_df, age_graph=age_graph, category_graph=category_graph,
                                       store_graph=store_graph, potential_sales=potential_sales, total_potential_customers=total_potential_customers,
                                       eating_places_df=eating_places_df, restaurant_type_df=restaurant_type_df,
                                       restaurant_summary_graph=restaurant_summary_graph, restaurants_graph=restaurants_graph)

            except SheetErrors.SheetErrors as e:
                flash(e.message + ". Please recheck the Spreadsheet and make sure your tables have the appropriate Named ranges and try again!")
                return redirect("/")

        flash(u'Your URL was not a valid Sheets Spreadsheet URL. Please try again.', 'error')
        return redirect("/")
        '''
    return render_template("home.jinja2")

@plot_blueprint.route('/e_plots', methods=['POST','GET'])
def create_eplots():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file part.')
            return redirect(request.url)
        file = request.files['file']

        if file.filename == '':
            flash('No selected file.')
            return redirect(request.url)
        if file and allowed_file(file.filename):
            filename = "for file " + secure_filename(file.filename)
            wb = load_workbook(file)
        #print(type(wb))

        try:
            x = return_prepped_dfs_from_excel(wb)
            ages_df = x[0]
            chain_df = x[1]
            stores_df = x[2]
            eating_places_df = x[3]
            restaurant_type_df = x[4]

            age_graph = create_age_plot(ages_df)
            category_graph = create_category_plot(chain_df)
            store_graph = create_stores_plot(stores_df)
            restaurant_summary_graph = create_restaurant_summary_plot(restaurant_type_df)
            restaurants_graph = create_restaurant_plot(eating_places_df)

            potential_sales = potential_sales_df(stores_df, chain_df)
            total_potential_customers = potential_sales["Potential Additional Customers"].sum()

            return render_template('plots/plots.jinja2', chain_df=chain_df, stores_df=stores_df, ages_df=ages_df, age_graph=age_graph, category_graph=category_graph,
                                   store_graph=store_graph, filename=filename, potential_sales=potential_sales,
                                   total_potential_customers=total_potential_customers, eating_places_df=eating_places_df, restaurant_type_df=restaurant_type_df,
                                   restaurant_summary_graph=restaurant_summary_graph, restaurants_graph=restaurants_graph)
        except Exception as e:
            if 'None' in str(e):
                flash('Please make sure to convert any tables into ranges!')
            else:
                flash(str(e))
            return render_template("home.jinja2")

    return render_template("home.jinja2")

@plot_blueprint.route('/create_static', methods=['POST','GET'])
def create_static():
    if request.method == 'POST':
        current_file_directory = os.path.dirname(__file__)
        excel_FP = os.path.join(current_file_directory, 'inMarket_Data_Challenge.xlsx')

        filename = "inMarket Data Challenge #1"
        wb = load_workbook(excel_FP)

        try:
            x = return_prepped_dfs_from_excel(wb)
            ages_df = x[0]
            chain_df = x[1]
            stores_df = x[2]
            eating_places_df = x[3]
            restaurant_type_df = x[4]

            age_graph = create_age_plot(ages_df)
            category_graph = create_category_plot(chain_df)
            store_graph = create_stores_plot(stores_df)
            restaurant_summary_graph = create_restaurant_summary_plot(restaurant_type_df)
            restaurants_graph = create_restaurant_plot(eating_places_df)

            potential_sales = potential_sales_df(stores_df, chain_df)
            total_potential_customers = potential_sales["Potential Additional Customers"].sum()


            return render_template('plots/plots.jinja2', chain_df=chain_df, stores_df=stores_df, ages_df=ages_df, age_graph=age_graph, category_graph=category_graph,
                                   store_graph=store_graph, filename=filename, potential_sales=potential_sales,
                                   total_potential_customers=total_potential_customers, eating_places_df=eating_places_df, restaurant_type_df=restaurant_type_df,
                                   restaurant_summary_graph=restaurant_summary_graph, restaurants_graph=restaurants_graph)
        except Exception as e:
            if 'None' in str(e):
                flash('Please make sure to convert any tables into ranges!')
            else:
                flash(str(e))
            return render_template("home.jinja2")

    return render_template("home.jinja2")




def allowed_file(filename):
    return '.' in filename and \
        filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def create_age_plot(ages_df):
    col_list = ['Customer_Buy_%', '% of Total Sales', '% of Total Foot Traffic']
    val_list = []

    for age in ages_df.index.values:
        temp_list = []
        temp_list.append(age)
        for col in col_list:
            temp_list.append(ages_df.loc[age, col])
        val_list.append(temp_list)

    try:
        age_graph = pygal.XY(stroke=False, title='Age Range Plot',legend_at_bottom = True,x_title='Customer Buy %', y_title='% of Total Sales',style=DarkStyle)
        age_graph.x_labels = [0.00, 0.25, 0.5, 0.75, 1.00]
        age_graph.y_labels = [0.0,0.1,0.2,0.3,0.4,0.5]
        for row in val_list:
            age_graph.add(row[0], [{'value':[row[1], row[2]],'label':'% of Total Population = {}'.format(row[3])}], dots_size=row[3]*100)
        graph_data = age_graph.render_data_uri()
        return graph_data

    except Exception as e:
        return(str(e))


def create_category_plot(chain_df):

    col_list = ['Customer_Buy_%', '% of Total Sales', 'Average # of Sales per Store']
    val_list = []

    for category in chain_df.index.values:
        temp_list = []
        temp_list.append(category)
        for col in col_list:
            temp_list.append(chain_df.loc[category, col])
        val_list.append(temp_list)

    try:
        category_graph = pygal.XY(stroke=False,title='Categories Plot', legend_at_bottom = True, x_title='Customer Buy %', y_title="% of Total Sales",style=DarkStyle)
        category_graph.x_labels = [0.30,0.40,0.50,0.60]
        category_graph.y_labels = [0.0,0.25,0.5,0.80]
        for row in val_list:
            category_graph.add(row[0], [{'value':[row[1], row[2]],'label':'Average # of Sales/Store = {}'.format(row[3])}], dots_size=(row[3] / 30))
        graph_data = category_graph.render_data_uri()
        return graph_data
    except Exception as e:
        return str(e)


def create_stores_plot(stores_df):
    col_list = ['Chain', 'Customer_Buy_%', '% of Total Sales', 'Total Foot Traffic', 'Traffic Score']
    cat_list = list(stores_df['Chain Category'].unique())
    val_dict = {}

    """
    val_dict = {
        "Drinking": [["Subway",0.54,0.76,11000, 0.912],[]],
        "Eating Places": [[]]
    }
    """

    for category in cat_list:
        cat_df = stores_df[stores_df['Chain Category'] == category].reset_index()
        store_list = []
        for index, row in cat_df.iterrows():
            store_row = []
            for col in col_list:
                store_row.append(row[col])
            store_list.append(store_row)
        val_dict[category] = store_list

    try:
        store_graph = pygal.XY(
            stroke=False,title="Stores Plot",
            x_title="Customer Buy %",
            y_title='Traffic Score',
            style=DarkStyle,
            legend_at_bottom=True
        )

        for cat in cat_list:
            # graph.add(cat, [])
            temp_list = []
            for store in val_dict[cat]:
                temp_list.append({'value': (store[1], store[4]), 'label': store[0]})
            store_graph.add(
                cat, temp_list, dots_size=5
            )

        graph_data = store_graph.render_data_uri()
        return graph_data
    except Exception as e:
        return(str(e))

def create_restaurant_summary_plot(restaurant_type_df):
    col_list = ['Average Customer Buy %', '% of Total Sales within Eating Places', '# of Restaurants']
    val_list = []

    for category in restaurant_type_df.index.values:
        temp_list = []
        temp_list.append(category)
        for col in col_list:
            temp_list.append(restaurant_type_df.loc[category, col])
        val_list.append(temp_list)

    try:
        category_graph = pygal.XY(stroke=False, title='Restaurants Summary by Type', legend_at_bottom=True,
                                  x_title='Average Customer Buy %', y_title="% of Total Sales within Category", style=DarkStyle)
        category_graph.x_labels = [0.30, 0.40, 0.50, 0.60]
        category_graph.y_labels = [0.0, 0.25, 0.5, 0.80]
        for row in val_list:
            category_graph.add(row[0],
                               [{'value': [row[1], row[2]], 'label': 'Total # of Restaurants = {}'.format(row[3])}],
                               dots_size=(row[3]/2))
        graph_data = category_graph.render_data_uri()
        return graph_data
    except Exception as e:
        return str(e)

def create_restaurant_plot(eating_places_df):
    col_list = ['Chain', 'Customer_Buy_%', '% of Total Sales within Eating Places']
    cat_list = list(eating_places_df['Restaurant Type'].unique())
    val_dict = {}

    """
    val_dict = {
        "Drinking": [["Subway",0.54,0.76,11000, 0.912],[]],
        "Eating Places": [[]]
    }
    """

    for category in cat_list:
        cat_df = eating_places_df[eating_places_df['Restaurant Type'] == category].reset_index()
        store_list = []
        for index, row in cat_df.iterrows():
            store_row = []
            for col in col_list:
                store_row.append(row[col])
            store_list.append(store_row)
        val_dict[category] = store_list

    try:
        store_graph = pygal.XY(
            stroke=False, title="Eating Places Plot",
            x_title="Customer Buy %",
            y_title='% of Total Sales within Category',
            style=DarkStyle,
            legend_at_bottom=True
        )

        for cat in cat_list:
            temp_list = []
            for store in val_dict[cat]:
                temp_list.append({'value': (store[1], store[2]), 'label': store[0]})
            store_graph.add(
                cat, temp_list, dots_size=5
            )

        graph_data = store_graph.render_data_uri()
        return graph_data
    except Exception as e:
        return str(e)














